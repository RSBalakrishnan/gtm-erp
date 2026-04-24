import openpyxl
import json
import os
import sys

def ingest_excel(excel_path, output_path):
    print(f"🕵️  Ingesting leads from {excel_path}...")
    
    if not os.path.exists(excel_path):
        print(f"❌ Error: {excel_path} not found.")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # Find the header row dynamically (within the first 20 rows)
    header_row_idx = None
    headers = []
    for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
        if row and "College Name" in row and "Website" in row:
            header_row_idx = i + 1  # 1-indexed for iter_rows
            headers = list(row)
            break
    
    if header_row_idx is None:
        print(f"❌ Error: Could not find header row with 'College Name' and 'Website'.")
        sys.exit(1)
        
    print(f"✅ Found headers at row {header_row_idx}.")
    try:
        name_idx = headers.index("College Name")
        url_idx = headers.index("Website")
        loc_idx = headers.index("Location")
    except ValueError as e:
        print(f"❌ Error: Missing required columns in headers: {headers}")
        sys.exit(1)
        
    leads = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row[url_idx]: continue
        
        leads.append({
            "college_name": row[name_idx],
            "website": row[url_idx],
            "location": row[loc_idx],
            "tracking": {
                "status": "open",
                "email_1_sent": False,
                "email_opened": False,
                "email_2_sent": False,
                "email_3_sent": False,
                "demo_booked": False
            }
        })
        
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(leads, f, indent=2)
        
    print(f"✅ Successfully ingested {len(leads)} leads to {output_path}.")

if __name__ == "__main__":
    EXCEL_FILE = "/Users/apple/Desktop/gtm/forscrapling.xlsx"
    OUTPUT_FILE = "/Users/apple/Desktop/gtm/agents/gtm-v2/workspace/leads.json"
    ingest_excel(EXCEL_FILE, OUTPUT_FILE)
